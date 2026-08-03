"""
changes_pdf.py — read the DEFRA "major changes" report (PDF) and retrieve the
passage relevant to a given factor. This is the grounding source for the AI
explanation: the model may only explain a change from text that actually appears
here. No vector DB — plain keyword + fuzzy retrieval, kept in memory.

Public functions:
    extract_changes_text(pdf_path) -> str
    chunk_changes(text)            -> list[dict{title, text}]
    load_change_chunks(pdf, xlsx)  -> list[dict{title, text, source, source_file}]
    retrieve_passage(chunks, material, min_score) -> (passage_text, score)
    retrieve_citation(chunks, material, min_score) -> dict | None

WHY TWO RETRIEVE FUNCTIONS. `retrieve_passage` returns the flat text the
explainer is grounded in, which is all the model ever needed. A cited memo needs
more: the section heading on its own, the quote on its own, and which document
the note came from, so a reader can check the claim instead of trusting a badge.
`retrieve_citation` returns that, and both call the SAME `_best_chunk`, so the
two can never disagree about which passage won.
"""

from __future__ import annotations

import os
import re

import pdfplumber
import openpyxl
from rapidfuzz import fuzz

# Generic words that carry no discriminating signal when matching a material name
# to a passage (they appear on almost every DEFRA row).
_STOPWORDS = {
    "primary", "material", "production", "average", "blend", "generated",
    "supply", "scope", "the", "and", "of", "for", "uk", "all", "use", "mix",
    "total", "per", "unit", "factor", "factors",
}


def extract_changes_text(pdf_path: str) -> str:
    """Extract all text from the changes-report PDF."""
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def chunk_changes(text: str) -> list[dict]:
    """
    Split the report into passages. We prefer explicit "Change: <x>" blocks (as
    DEFRA-style reports use headed sections); if none exist we fall back to
    paragraph blocks separated by blank lines.
    """
    lines = [ln.rstrip() for ln in text.splitlines()]

    # Try headed "Change:" blocks first.
    blocks, current = [], None
    for ln in lines:
        m = re.match(r"\s*change\s*:\s*(.+)", ln, flags=re.IGNORECASE)
        if m:
            if current:
                blocks.append(current)
            current = {"title": m.group(1).strip(), "text": ""}
        elif current is not None:
            current["text"] = (current["text"] + " " + ln).strip()
    if current:
        blocks.append(current)

    if blocks:
        return [b for b in blocks if b["text"] or b["title"]]

    # Fallback: paragraph blocks.
    paras, buf = [], []
    for ln in lines:
        if ln.strip():
            buf.append(ln.strip())
        elif buf:
            paras.append(" ".join(buf))
            buf = []
    if buf:
        paras.append(" ".join(buf))
    return [{"title": "", "text": p} for p in paras if len(p) > 20]


def extract_whats_new_text(xlsx_path: str) -> str:
    """
    Extract the "What's new" sheet from a DEFRA workbook. The real full-set
    workbook ships this sheet, and it is an excellent grounding source: it
    explains the year's methodology revisions, data updates and relabels in
    prose. Used when a standalone Major Changes PDF is not provided.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    target = None
    for name in wb.sheetnames:
        if "what" in name.lower() and "new" in name.lower():
            target = name
            break
    if target is None:
        return ""
    ws = wb[target]
    rows = []
    for row in ws.iter_rows(values_only=True):
        text = " ".join(str(c) for c in row if c is not None and str(c).strip())
        if text.strip():
            rows.append(text.strip())
    return "\n".join(rows)


def chunk_whats_new(text: str) -> list[dict]:
    """
    Chunk the "What's new" text by its numbered sections, e.g.
    "1 Revision to the calculation method for UK electricity". Each section
    becomes one retrievable passage.
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    blocks, current = [], None
    heading = re.compile(r"^(\d+)\s+([A-Z].{6,})")
    for ln in lines:
        m = heading.match(ln)
        if m:
            if current:
                blocks.append(current)
            current = {"title": m.group(2).strip(), "text": ""}
        elif current is not None:
            current["text"] = (current["text"] + " " + ln).strip()
    if current:
        blocks.append(current)
    return [b for b in blocks if b["text"]]


def load_change_chunks(pdf_path: str | None, new_workbook_path: str | None) -> list[dict]:
    """
    Build grounding chunks, preferring a real Major Changes PDF, then falling
    back to the new workbook's "What's new" sheet. Returns [] if neither works
    (the explainer then honestly reports "no official reason found").

    Each chunk is TAGGED with the document it came from. Without the tag a memo
    could only say "the DEFRA notes", which is not a citation a reader can follow:
    the PDF and the workbook sheet are different documents, and which one was read
    is part of the claim.
    """
    if pdf_path:
        try:
            chunks = chunk_changes(extract_changes_text(pdf_path))
            if chunks:
                return _tag(chunks, "Major Changes report", pdf_path)
        except Exception:
            pass
    if new_workbook_path:
        try:
            chunks = chunk_whats_new(extract_whats_new_text(new_workbook_path))
            return _tag(chunks, '"What\'s new" sheet', new_workbook_path)
        except Exception:
            pass
    return []


def _tag(chunks: list[dict], source: str, path: str) -> list[dict]:
    """Stamp every chunk with the document it was read from."""
    name = os.path.basename(path)
    for c in chunks:
        c["source"] = source
        c["source_file"] = name
    return chunks


def _keywords(material: str) -> set[str]:
    """Discriminating tokens from a material/activity name."""
    toks = re.findall(r"[a-z0-9]+", str(material).lower())
    return {t for t in toks if len(t) >= 4 and t not in _STOPWORDS}


def _keyword_overlap(material: str, passage: str) -> float:
    """Fraction of the material's discriminating keywords present in the passage."""
    kws = _keywords(material)
    if not kws:
        return 0.0
    ptoks = set(re.findall(r"[a-z0-9]+", passage.lower()))
    hit = sum(1 for k in kws if k in ptoks)
    return hit / len(kws)


def _best_chunk(chunks: list[dict], material: str, min_score: float = 0.5):
    """
    The single ranking function. Returns (chunk, score), or (None, score) when
    nothing clears the bar. A returned None is the signal that the report does
    NOT explain this change, and the explainer must then say so rather than
    invent a reason.

    DECISIONS D11 lives in this function. Do not change the scoring here to make
    a citation render more often: a citation that names the wrong note is worse
    than no citation at all.
    """
    best_chunk, best_score = None, 0.0
    for ch in chunks:
        haystack = f"{ch['title']} {ch['text']}"
        overlap = _keyword_overlap(material, haystack)
        # Fuzzy score against the (short) title refines the ranking, but only for
        # a passage that ALREADY shares the query's distinctive keywords. On its
        # own, title fuzz matches shared boilerplate: "Petrol (average biofuel
        # blend)" scores ~0.87 against the DIESEL note's title because they share
        # "(average biofuel blend)", and a generic "Calculating emissions"
        # heading scores ~0.55 against many factor names. Letting title fuzz
        # trigger a hit by itself grounds a change in the WRONG note, which is the
        # one failure the tool must never make. So keyword overlap is the gate:
        # the title can only boost a passage whose overlap already clears the bar.
        title_fuzz = fuzz.token_set_ratio(material.lower(), ch["title"].lower()) / 100.0
        score = max(overlap, title_fuzz) if overlap >= min_score else overlap
        if score > best_score:
            best_score, best_chunk = score, ch

    if best_score >= min_score:
        return best_chunk, round(best_score, 3)
    return None, round(best_score, 3)


def retrieve_passage(chunks: list[dict], material: str, min_score: float = 0.5):
    """
    Return (passage_text, score) for the chunk best matching `material`, or
    ("", 0.0) if nothing is relevant enough. This is the text the explainer is
    grounded in, and its shape is unchanged: heading and body joined, because
    that is what the model reads.
    """
    ch, score = _best_chunk(chunks, material, min_score)
    if ch is None:
        return "", score
    return f"{ch['title']} {ch['text']}".strip(), score


def retrieve_citation(chunks: list[dict], material: str, min_score: float = 0.5):
    """
    The same winning chunk, kept in pieces so it can be CITED rather than just
    read: heading on its own, quote on its own, and the document both came from.

    Returns None when nothing clears the bar, which is the same signal
    `retrieve_passage` gives by returning "". Both call `_best_chunk`, so the
    quote shown to a reader is always the passage the explanation was built from.
    """
    ch, score = _best_chunk(chunks, material, min_score)
    if ch is None:
        return None
    return {
        "heading": (ch.get("title") or "").strip(),
        "quote": (ch.get("text") or "").strip(),
        # Absent only for chunks built directly by a test or a caller that
        # bypassed load_change_chunks. Stated, never invented (DECISIONS D2).
        "source": ch.get("source", "source not recorded"),
        "source_file": ch.get("source_file", ""),
        "score": score,
    }
