"""
test_export_pack.py - what the consultant actually hands over.

The export is the deliverable. If it is wrong, or if it quietly looks cleaner
than the run really was, the tool has done damage rather than work. Three things
are defended here:

1. FOUR ARTIFACTS, ONE RUN. The workbook, the JSON, the Markdown and the memo all
   carry the same run id, so a page pulled out of a client folder can be traced
   back to the run that produced it.
2. NOTHING SHIPS SILENTLY. Whatever the completeness checklist leaves unresolved
   appears in the front matter of every artifact. A run with lines held for
   review cannot produce a document that reads as complete.
3. THE MEMO IS SELF-CONTAINED AND PRINTABLE. No external requests, real tables,
   nothing collapsed, headers repeated across pages.

Run it on its own to SEE the memo:

    python tests/test_export_pack.py     # writes the memo to the scratch dir

"""

from __future__ import annotations

import io
import json
import os
import sys
import zipfile
from datetime import datetime, timezone

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "src"))

import export as ex                     # noqa: E402
from explain import NO_REASON           # noqa: E402
from pipeline import run_pipeline       # noqa: E402

SYNTH = os.path.join(ROOT, "data", "synthetic")
FIXED_TIME = datetime(2026, 7, 31, 9, 30, 0, tzinfo=timezone.utc)


@pytest.fixture(scope="module")
def results():
    return run_pipeline(
        defra_old_path=os.path.join(SYNTH, "defra_2025.xlsx"),
        defra_new_path=os.path.join(SYNTH, "defra_2026.xlsx"),
        changes_pdf_path=os.path.join(SYNTH, "changes_2026.pdf"),
        bom_path_or_df=os.path.join(SYNTH, "sample_bom.csv"),
        old_label="2025",
        new_label="2026",
    )


@pytest.fixture(scope="module")
def identity(results):
    return ex.run_identity(
        results,
        client="Acme Ltd",
        product="Widget",
        operator="S. Pala",
        generated_utc=FIXED_TIME,
    )


@pytest.fixture(scope="module")
def checklist(results):
    return ex.completeness_checklist(results, set_aside=[{"row_number": 4}])


@pytest.fixture(scope="module")
def pack(results, identity, checklist):
    return ex.export_pack(results, identity, checklist)


# ---------------------------------------------------------------------------
# Identity
# ---------------------------------------------------------------------------

def test_the_run_id_is_the_same_in_all_four_artifacts(pack, identity):
    run_id = identity["run_id"]
    assert len(pack) == 4

    assert run_id in pack[[n for n in pack if n.endswith(".md")][0]]
    assert run_id in pack[[n for n in pack if n.endswith(".json")][0]]
    assert run_id in pack[[n for n in pack if n.endswith(".html")][0]]

    workbook = _open_workbook(pack)
    for sheet in workbook.worksheets:
        assert run_id in str(sheet.cell(row=1, column=1).value), (
            f"sheet {sheet.title} does not carry the run id"
        )


def test_the_run_id_is_hashed_from_the_inputs_not_the_clock(results):
    """Re-running the same analysis gives the same id, so a changed id means
    something about the run genuinely changed."""
    first = ex.run_identity(results, client="Acme Ltd", generated_utc=FIXED_TIME)
    later = ex.run_identity(
        results, client="Acme Ltd", generated_utc=datetime(2027, 1, 1, tzinfo=timezone.utc)
    )
    assert first["run_id"] == later["run_id"]
    assert first["generated_utc"] != later["generated_utc"]

    other_client = ex.run_identity(results, client="Beta Ltd", generated_utc=FIXED_TIME)
    assert other_client["run_id"] != first["run_id"]


def test_the_mapping_fingerprint_tracks_which_line_became_which_activity(results):
    """The thing worth comparing between years."""
    first = ex.mapping_hash(results["matched_df"])

    changed = results["matched_df"].copy()
    changed.loc[changed.index[0], "matched_activity"] = "Something else entirely"
    assert ex.mapping_hash(changed) != first


def test_nothing_is_written_to_disk(results, identity, checklist, tmp_path):
    """The no-database rule: the pack is built in memory and handed back."""
    before = set(os.listdir(ROOT))
    ex.export_pack(results, identity, checklist)
    assert set(os.listdir(ROOT)) == before


# ---------------------------------------------------------------------------
# Completeness
# ---------------------------------------------------------------------------

def test_the_checklist_reports_what_is_open(results):
    items = ex.completeness_checklist(results, set_aside=[{"row_number": 4}])
    keys = {item["key"] for item in items}
    assert keys == {
        "every_line_matched", "no_rows_set_aside", "coverage_at_bar", "changes_are_cited",
    }

    set_aside_item = next(i for i in items if i["key"] == "no_rows_set_aside")
    assert set_aside_item["resolved"] is False
    assert "1 row(s)" in set_aside_item["detail"]


def test_a_run_with_nothing_open_says_so(results):
    clean = ex.completeness_checklist(results, set_aside=[])
    item = next(i for i in clean if i["key"] == "no_rows_set_aside")
    assert item["resolved"] is True


def test_unresolved_items_reach_the_front_matter_of_every_artifact(pack, checklist):
    """The load-bearing promise: a consultant can always ship, never silently."""
    open_items = ex.unresolved(checklist)
    assert open_items, "the fixture run should have something open to check"
    first_label = open_items[0]["label"]

    markdown = pack[[n for n in pack if n.endswith(".md")][0]]
    assert "Unresolved at export" in markdown
    assert first_label in markdown

    memo = pack[[n for n in pack if n.endswith(".html")][0]]
    assert "Unresolved at export" in memo
    assert first_label in memo

    payload = json.loads(pack[[n for n in pack if n.endswith(".json")][0]])
    assert first_label in payload["unresolved"]

    workbook = _open_workbook(pack)
    method = workbook["Method"]
    text = "\n".join(str(cell.value) for row in method.iter_rows() for cell in row)
    assert "UNRESOLVED" in text


def test_an_unexplained_change_is_reported_as_unexplained_not_hidden(results):
    """The domain rule, at export: silence in the DEFRA notes is stated, not filled."""
    items = ex.completeness_checklist(results)
    cited = next(i for i in items if i["key"] == "changes_are_cited")
    unexplained = [
        e for e in results["explanations"]
        if e["plain_english_reason"].strip() == NO_REASON
    ]
    assert cited["resolved"] == (not unexplained)


# ---------------------------------------------------------------------------
# The workbook
# ---------------------------------------------------------------------------

def test_the_workbook_carries_movers_mapping_review_and_method(pack):
    workbook = _open_workbook(pack)
    assert workbook.sheetnames == ["Movers", "Mapping", "Review", "Method"]


def test_the_method_sheet_records_how_the_run_was_configured(pack, identity):
    workbook = _open_workbook(pack)
    text = "\n".join(
        str(cell.value) for row in workbook["Method"].iter_rows() for cell in row
    )
    assert identity["mapping_hash"] in text
    assert identity["version_old"] in text
    assert identity["version_new"] in text


# ---------------------------------------------------------------------------
# The memo
# ---------------------------------------------------------------------------

def test_the_memo_makes_no_external_requests(pack):
    """A document holding a client's confidential inventory should not phone out."""
    memo = pack[[n for n in pack if n.endswith(".html")][0]]
    for probe in ("http://", "https://", "<script", "@import", "//fonts."):
        assert probe not in memo, f"the memo reaches outside itself: {probe}"


def test_the_memo_is_a_real_document_with_real_tables(pack):
    memo = pack[[n for n in pack if n.endswith(".html")][0]]
    assert memo.startswith("<!doctype html>")
    assert "<table>" in memo and "<caption>" in memo and 'scope="col"' in memo
    assert "<h1>" in memo and "<h2>" in memo
    assert "display: table-header-group" in memo, "printed pages must repeat headers"
    assert "@media print" in memo


def test_the_memo_hides_nothing_behind_a_disclosure(pack):
    """A printed report must contain the reasoning, not a row of closed drawers."""
    memo = pack[[n for n in pack if n.endswith(".html")][0]]
    assert "<details" not in memo


def test_the_memo_states_direction_in_words(pack):
    memo = pack[[n for n in pack if n.endswith(".html")][0]]
    assert "visually-hidden" in memo
    assert any(f", {word}</span>" in memo for word in ("rose", "fell", "did not change"))


def test_the_memo_escapes_a_hostile_line_item(results, identity, checklist):
    """Line items come from an uploaded file and end up in a shared document.

    The payload is planted in both places the memo prints a line item, the
    review table and the movers table, and each is asserted to arrive ESCAPED
    rather than merely absent, so the test cannot pass by rendering nothing.
    """
    payload = "<script>alert(1)</script>"
    hostile = dict(results)

    matched = results["matched_df"].copy()
    review_rows = matched.index[matched["needs_review"]]
    assert len(review_rows), "fixture should hold at least one line for review"
    matched.loc[review_rows[0], "line_item"] = payload
    hostile["matched_df"] = matched

    top = results["top_delta"].copy()
    assert len(top), "fixture should have at least one mover"
    top.loc[top.index[0], "line_item"] = payload
    hostile["top_delta"] = top

    memo = ex.to_print_html(hostile, identity, checklist)
    assert payload not in memo
    assert memo.count("&lt;script&gt;alert(1)&lt;/script&gt;") == 2


# ---------------------------------------------------------------------------
# The zip
# ---------------------------------------------------------------------------

def test_the_pack_downloads_as_one_archive(pack):
    archive = zipfile.ZipFile(io.BytesIO(ex.to_zip(pack)))
    assert sorted(archive.namelist()) == sorted(pack)
    assert not archive.testzip()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _open_workbook(pack):
    import openpyxl

    name = [n for n in pack if n.endswith(".xlsx")][0]
    return openpyxl.load_workbook(io.BytesIO(pack[name]))


# ---------------------------------------------------------------------------
# Seeing it work
# ---------------------------------------------------------------------------

def _demo():
    built = run_pipeline(
        defra_old_path=os.path.join(SYNTH, "defra_2025.xlsx"),
        defra_new_path=os.path.join(SYNTH, "defra_2026.xlsx"),
        changes_pdf_path=os.path.join(SYNTH, "changes_2026.pdf"),
        bom_path_or_df=os.path.join(SYNTH, "sample_bom.csv"),
        old_label="2025",
        new_label="2026",
    )
    ident = ex.run_identity(built, client="Acme Ltd", product="Widget", operator="S. Pala")
    check = ex.completeness_checklist(built)
    files = ex.export_pack(built, ident, check)

    out = os.path.join(ROOT, "reports")
    os.makedirs(out, exist_ok=True)
    for name, content in files.items():
        path = os.path.join(out, name)
        mode, payload = ("wb", content) if isinstance(content, bytes) else ("w", content)
        with open(path, mode, **({} if mode == "wb" else {"encoding": "utf-8"})) as fh:
            fh.write(payload)
        print(f"{len(content):>9,} bytes  {path}")

    print(f"\nRun id {ident['run_id']}, generated {ident['generated_utc']}")
    for item in check:
        print(f"  [{'x' if item['resolved'] else ' '}] {item['label']}. {item['detail']}")
    print("\nOpen the .html in a browser and use Print to PDF.")
    return 0


if __name__ == "__main__":
    raise SystemExit(_demo())
